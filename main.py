# -*- coding: utf-8 -*-
"""
Microservicio Contaser - Módulo 1: COMPRAS FC ELEC (v1.1)
Lee un .xlsm, calcula la hoja "COMPRAS FC ELEC" y escribe la tabla de
resultados mediante EDICIÓN QUIRÚRGICA del XML interno: solo se modifica
el XML de esa hoja; el resto del archivo (58 hojas, macros, dibujos,
vínculos externos) se copia byte por byte, evitando corrupción.
"""
import base64
import io
import os
import re
import unicodedata
import zipfile
from datetime import datetime, date
from xml.sax.saxutils import escape, unescape

from fastapi import FastAPI, UploadFile, File, Form, Header, HTTPException
import openpyxl
from openpyxl.utils import get_column_letter

app = FastAPI(title="Contaser - COMPRAS FC ELEC", version="1.9")

SERVICE_TOKEN = os.getenv("SERVICE_TOKEN", "")  # si está vacío, no exige token
SHEET_NAME = "COMPRAS FC ELEC"
END_MARKER = "facturas procesadas disponibles"
# Título de la tabla que insertamos. Sirve además como marcador de idempotencia:
# si ya existe en la hoja, se borra la tabla anterior antes de escribir la nueva.
MARCADOR_TABLA = "RESULTADOS COMPRAS FC ELEC"

COL_LABELS = {
    "identificación emisor factura": "nit_emisor",
    "nombre emisor factura": "nombre_emisor",
    "fecha emisión": "fecha_emision",
    "valor facturado": "valor_facturado",
    "valor notas crédito": "notas_credito",
    "valor notas débito": "notas_debito",
    "valor factura / afectada con notas débito - crédito": "valor_neto",
    "valor susceptible beneficio": "valor_beneficio",
    "medios de pago": "medio_pago",
    "num_factura_venta": "num_factura",
}


def _norm(s):
    return str(s).strip().lower() if s is not None else ""


def _sin_tildes(s):
    return "".join(c for c in unicodedata.normalize("NFD", str(s))
                   if unicodedata.category(c) != "Mn")


def _norm_label(s):
    """Normaliza una etiqueta del header para compararla con tolerancia.

    Quita tildes, pasa a minúsculas, convierte los dos puntos en espacio y
    colapsa espacios múltiples. Necesario porque el export de la DIAN cambia
    de formato entre versiones: 'Nombre' en unos archivos y 'Nombre o razón
    social' en otros, 'Identificacion ' con espacio final, etc.
    """
    t = _sin_tildes(s or "").lower().replace(":", " ")
    return re.sub(r"\s+", " ", t).strip()


def nombre_desde_archivo(nombre_archivo):
    """Fallback del nombre de cliente: usa el nombre del archivo sin extensión
    ni el sufijo ' - AAAA'. Devuelve '' si no se puede derivar nada."""
    if not nombre_archivo:
        return ""
    base = re.sub(r"\.(xlsm|xlsx)$", "", str(nombre_archivo), flags=re.I)
    base = re.sub(r"\s*-\s*\d{4}\s*$", "", base)
    return base.strip()


def to_number(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if s in ("", "-"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def to_iso(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip() if v is not None else ""


def classify_medio(v):
    s = _norm(v)
    if s.startswith("efectivo y"):
        return "efectivo_otros"
    if "lectr" in s:
        return "electronico"
    if "fectivo" in s:
        return "efectivo"
    return "otro"


def _norm_title(s):
    """Normaliza título de hoja: minúsculas, colapsa espacios múltiples."""
    return re.sub(r"\s+", " ", str(s)).strip().lower() if s is not None else ""


def find_sheet(wb):
    objetivo = _norm_title(SHEET_NAME)
    for ws in wb.worksheets:
        if _norm_title(ws.title) == objetivo:
            return ws
    for ws in wb.worksheets:
        if objetivo in _norm_title(ws.title):
            return ws
    # variante sin espacios (p.ej. "COMPRASFCELEC")
    compacto = objetivo.replace(" ", "")
    for ws in wb.worksheets:
        if compacto in _norm_title(ws.title).replace(" ", ""):
            return ws
    hojas = ", ".join(repr(ws.title) for ws in wb.worksheets)
    raise HTTPException(400, f"No se encontró la hoja '{SHEET_NAME}'. Hojas del archivo: [{hojas}]")


# Márgenes de barrido. Las columnas de la DIAN pueden correrse de posición si
# se agregan columnas nuevas; los nombres no cambian, pero la posición sí.
MAX_FILA_HEADER = 80      # hasta dónde se busca la fila de encabezados
MAX_COL_HEADER = 40       # hasta qué columna se buscan las etiquetas


def find_header_row(ws):
    for r in range(1, MAX_FILA_HEADER):
        for c in range(1, MAX_COL_HEADER):
            if _norm(ws.cell(row=r, column=c).value) == "identificación emisor factura":
                return r
    raise HTTPException(400, "No se encontró la fila de encabezados de datos (COMPRAS FC ELEC)")


def map_columns(ws, header_row):
    colmap = {}
    encontradas = []
    for c in range(1, MAX_COL_HEADER + 1):
        crudo = ws.cell(row=header_row, column=c).value
        label = _norm(crudo)
        if not label:
            continue
        encontradas.append(f"{get_column_letter(c)}={crudo}")
        if label.startswith("cufe"):
            colmap["cufe"] = c
        elif label in COL_LABELS:
            colmap[COL_LABELS[label]] = c
    requeridas = ["nit_emisor", "valor_facturado", "valor_neto", "valor_beneficio", "medio_pago", "cufe"]
    faltan = [k for k in requeridas if k not in colmap]
    if faltan:
        # Incluir los encabezados reales: sin esto el diagnóstico es a ciegas.
        raise HTTPException(
            400, f"Faltan columnas en el encabezado: {faltan}. "
                 f"Encabezados encontrados en la fila {header_row}: {encontradas[:25]}")
    return colmap


def _celda(ws, fila, colmap, campo):
    """Valor de una columna mapeada, o None si esa columna no existe.

    Solo 6 de las 11 columnas son obligatorias; sin esto, un archivo sin
    'Valor Notas Crédito' reventaba con KeyError y un 500 sin explicación.
    """
    c = colmap.get(campo)
    return ws.cell(row=fila, column=c).value if c else None


def parse_header(ws, header_row, nombre_archivo=None):
    """Extrae año gravable, NIT y nombre de las filas previas al encabezado.

    Las etiquetas se comparan con `_norm_label` y por prefijo, no por igualdad:
    'Nombre', 'Nombre o razón social' y 'Nombres y apellidos' deben funcionar
    todas. Si aun así el nombre queda vacío, cae al nombre del archivo.
    """
    anio = nit = nombre = None
    nit_row = None
    for r in range(1, header_row):
        label = valor = None
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c).value
            if cell is not None and str(cell).strip() != "":
                if label is None:
                    label = _norm_label(cell)
                else:
                    valor = cell
                    break
        if not label:
            continue
        if label.startswith("ano gravable"):        # 'Año Gravable' sin tilde
            anio = valor
        elif label == "nit":
            nit = valor
            nit_row = r
        elif label.startswith("nombre"):
            # La fila 1 trae label 'Nombre' con el título del informe: descartar.
            if valor is None or _norm_label(valor).startswith("informe"):
                continue
            # Preferir la fila posterior al NIT; si no hay, la primera válida.
            if nit_row is not None and r > nit_row:
                nombre = valor
            elif nombre is None:
                nombre = valor

    if nombre is None or str(nombre).strip() == "":
        nombre = nombre_desde_archivo(nombre_archivo) or None
    return anio, nit, nombre


# ---------------------------------------------------------------------------
# Escritura quirúrgica del XML de la hoja (sin reescribir el resto del libro)
# ---------------------------------------------------------------------------

def _sheet_xml_path(z: zipfile.ZipFile, sheet_title: str) -> str:
    wb_xml = z.read("xl/workbook.xml").decode("utf-8")
    rid = None
    for sm in re.finditer(r"<sheet\b[^>]*>", wb_xml):
        tag = sm.group(0)
        nm = re.search(r'name="([^"]*)"', tag)
        if nm and unescape(nm.group(1)) == sheet_title:
            rm = re.search(r'r:id="([^"]*)"', tag)
            if rm:
                rid = rm.group(1)
            break
    if not rid:
        raise HTTPException(500, f"No se encontró la hoja '{sheet_title}' en workbook.xml")
    rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    rel = re.search(r'<Relationship\b[^>]*\bId="%s"[^>]*/?>' % re.escape(rid), rels)
    if not rel:
        raise HTTPException(500, f"No se encontró la relación {rid} de la hoja")
    tm = re.search(r'Target="([^"]*)"', rel.group(0))
    target = tm.group(1)
    return target[1:] if target.startswith("/") else "xl/" + target


def _bump_collection(xml: str, tag: str, items_xml: str, n_items: int):
    """Suma n_items al count de <tag> e inserta items antes de </tag>.
    Devuelve (xml_modificado, índice_inicial_de_los_nuevos_items)."""
    m = re.search(r'<%s count="(\d+)"' % tag, xml)
    if not m:
        raise HTTPException(500, f"styles.xml sin colección <{tag}>")
    old = int(m.group(1))
    xml = xml[:m.start(1)] + str(old + n_items) + xml[m.end(1):]
    close = xml.find("</%s>" % tag)
    if close == -1:
        raise HTTPException(500, f"styles.xml sin cierre </{tag}>")
    xml = xml[:close] + items_xml + xml[close:]
    return xml, old


# Paleta corporativa de la tabla de resultados (gama verde)
C_TITULO_FONDO = "FF375623"   # verde oscuro
C_HEADER_FONDO = "FF70AD47"   # verde medio
C_LABEL_FONDO = "FFE2EFDA"    # verde claro
C_TEXTO_OSCURO = "FF375623"


def _augment_styles(styles_xml: str) -> tuple:
    """Agrega fuentes/rellenos/bordes/formatos y 5 estilos de celda nuevos.
    Devuelve (styles_xml_nuevo, dict_estilos)."""
    # Formato de porcentaje 0.0% (id personalizado libre)
    ids = [int(x) for x in re.findall(r'numFmtId="(\d+)"', styles_xml)]
    pct_id = max([163] + ids) + 1
    numfmt = '<numFmt numFmtId="%d" formatCode="0.0%%"/>' % pct_id
    m = re.search(r'<numFmts count="(\d+)">', styles_xml)
    if m:
        styles_xml = styles_xml[:m.start(1)] + str(int(m.group(1)) + 1) + styles_xml[m.end(1):]
        close = styles_xml.find("</numFmts>")
        styles_xml = styles_xml[:close] + numfmt + styles_xml[close:]
    else:
        m2 = re.search(r"<styleSheet[^>]*>", styles_xml)
        styles_xml = (styles_xml[:m2.end()]
                      + '<numFmts count="1">%s</numFmts>' % numfmt
                      + styles_xml[m2.end():])

    fonts = (
        '<font><b/><sz val="12"/><color rgb="FFFFFFFF"/><name val="Calibri"/></font>'
        '<font><b/><sz val="11"/><color rgb="FFFFFFFF"/><name val="Calibri"/></font>'
        '<font><b/><sz val="11"/><color rgb="%s"/><name val="Calibri"/></font>' % C_TEXTO_OSCURO
    )
    styles_xml, f0 = _bump_collection(styles_xml, "fonts", fonts, 3)

    fills = (
        '<fill><patternFill patternType="solid"><fgColor rgb="%s"/><bgColor indexed="64"/></patternFill></fill>' % C_TITULO_FONDO
        + '<fill><patternFill patternType="solid"><fgColor rgb="%s"/><bgColor indexed="64"/></patternFill></fill>' % C_HEADER_FONDO
        + '<fill><patternFill patternType="solid"><fgColor rgb="%s"/><bgColor indexed="64"/></patternFill></fill>' % C_LABEL_FONDO
    )
    styles_xml, l0 = _bump_collection(styles_xml, "fills", fills, 3)

    border = ('<border><left style="thin"><color indexed="64"/></left>'
              '<right style="thin"><color indexed="64"/></right>'
              '<top style="thin"><color indexed="64"/></top>'
              '<bottom style="thin"><color indexed="64"/></bottom><diagonal/></border>')
    styles_xml, b0 = _bump_collection(styles_xml, "borders", border, 1)

    xfs = (
        # título (banda azul marino, texto blanco centrado)
        '<xf numFmtId="0" fontId="%d" fillId="%d" borderId="%d" xfId="0" applyFont="1" applyFill="1" applyBorder="1" applyAlignment="1"><alignment horizontal="center" vertical="center"/></xf>' % (f0, l0, b0)
        # encabezados Concepto/Valor (azul medio, texto blanco)
        + '<xf numFmtId="0" fontId="%d" fillId="%d" borderId="%d" xfId="0" applyFont="1" applyFill="1" applyBorder="1" applyAlignment="1"><alignment horizontal="center" vertical="center"/></xf>' % (f0 + 1, l0 + 1, b0)
        # etiquetas (azul claro, texto azul oscuro en negrilla)
        + '<xf numFmtId="0" fontId="%d" fillId="%d" borderId="%d" xfId="0" applyFont="1" applyFill="1" applyBorder="1" applyAlignment="1"><alignment horizontal="left" vertical="center" wrapText="1"/></xf>' % (f0 + 2, l0 + 2, b0)
        # valores numéricos con miles (#,##0 = formato incorporado 3)
        + '<xf numFmtId="3" fontId="0" fillId="0" borderId="%d" xfId="0" applyNumberFormat="1" applyBorder="1" applyAlignment="1"><alignment horizontal="right" vertical="center"/></xf>' % b0
        # porcentaje 0.0%%
        + '<xf numFmtId="%d" fontId="0" fillId="0" borderId="%d" xfId="0" applyNumberFormat="1" applyBorder="1" applyAlignment="1"><alignment horizontal="right" vertical="center"/></xf>' % (pct_id, b0)
    )
    styles_xml, x0 = _bump_collection(styles_xml, "cellXfs", xfs, 5)

    estilos = {"titulo": x0, "header": x0 + 1, "label": x0 + 2,
               "money": x0 + 3, "pct": x0 + 4}
    return styles_xml, estilos


def _indices_sharedstrings(z: zipfile.ZipFile, texto: str) -> set:
    """Índices de xl/sharedStrings.xml cuyo texto contiene `texto`.

    Nosotros escribimos la tabla con inlineStr, pero si alguien abre el archivo
    en Excel y lo guarda, Excel mueve esos textos a la tabla de cadenas
    compartidas y la celda pasa a ser <c t="s"><v>N</v></c>. Sin esto, la
    idempotencia no encontraría la tabla anterior y escribiría una segunda.
    """
    try:
        ss = z.read("xl/sharedStrings.xml").decode("utf-8", errors="replace")
    except KeyError:
        return set()
    encontrados = set()
    for i, m in enumerate(re.finditer(r"<si\b[^>]*>(.*?)</si>", ss, re.DOTALL)):
        plano = unescape(re.sub(r"<[^>]+>", "", m.group(1)))
        if texto in plano:
            encontrados.add(i)
    return encontrados


def write_results_table(content: bytes, sheet_title: str, last_row: int,
                        tabla: list) -> bytes:
    """Escribe la tabla de resultados en columnas B y C con título,
    encabezados, colores y formatos numéricos. tabla: [(label, value, kind)]
    con kind en {money, pct}."""
    zin = zipfile.ZipFile(io.BytesIO(content))
    sheet_path = _sheet_xml_path(zin, sheet_title)
    sxml = zin.read(sheet_path).decode("utf-8")
    styles_xml, st = _augment_styles(zin.read("xl/styles.xml").decode("utf-8"))

    col_l, col_v = "C", "D"

    # Idempotencia: si ya hay una tabla de resultados de una corrida anterior,
    # eliminarla (título + encabezados + 5 filas de datos) antes de escribir.
    row_re = re.compile(r'<row r="(\d+)"[^>]*?(?:/>|>.*?</row>)', re.DOTALL)
    sst_idx = _indices_sharedstrings(zin, MARCADOR_TABLA)
    cell_s_re = re.compile(r'<c\b[^>]*\bt="s"[^>]*>\s*<v>(\d+)</v>')
    title_row = None
    for m in row_re.finditer(sxml):
        bloque = m.group(0)
        # Caso 1: la tabla sigue como la escribimos (inlineStr).
        if MARCADOR_TABLA in bloque:
            title_row = int(m.group(1))
            break
        # Caso 2: Excel reguardó el archivo y el texto vive en sharedStrings.
        if sst_idx and any(int(cm.group(1)) in sst_idx
                           for cm in cell_s_re.finditer(bloque)):
            title_row = int(m.group(1))
            break
    if title_row is not None:
        def _drop(m):
            n = int(m.group(1))
            return "" if title_row <= n <= title_row + 6 else m.group(0)
        sxml = row_re.sub(_drop, sxml)

    existing = [int(x) for x in re.findall(r'<row r="(\d+)"', sxml)]
    max_existing = max(existing) if existing else last_row
    start = max(last_row + 3, max_existing + 2)

    def txt(col, r, s, text):
        return ('<c r="%s%d" s="%d" t="inlineStr"><is><t xml:space="preserve">%s</t></is></c>'
                % (col, r, s, escape(str(text))))

    rows_xml = []
    # Fila título (banda en B y C)
    rows_xml.append('<row r="%d" ht="20" customHeight="1">%s%s</row>' % (
        start,
        txt(col_l, start, st["titulo"], MARCADOR_TABLA),
        txt(col_v, start, st["titulo"], ""),
    ))
    # Fila encabezados
    rows_xml.append("<row r=\"%d\">%s%s</row>" % (
        start + 1,
        txt(col_l, start + 1, st["header"], "Concepto"),
        txt(col_v, start + 1, st["header"], "Valor"),
    ))
    # Filas de datos
    for i, (label, value, kind) in enumerate(tabla):
        r = start + 2 + i
        s_val = st["pct"] if kind == "pct" else st["money"]
        val = '<c r="%s%d" s="%d"><v>%s</v></c>' % (col_v, r, s_val, value)
        rows_xml.append('<row r="%d">%s%s</row>' % (
            r, txt(col_l, r, st["label"], label), val))

    idx = sxml.rfind("</sheetData>")
    if idx == -1:
        raise HTTPException(500, "XML de hoja sin </sheetData>; estructura inesperada")
    new_sxml = sxml[:idx] + "".join(rows_xml) + sxml[idx:]

    # Reempaquetar: todo idéntico excepto la hoja y styles.xml
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == sheet_path:
                data = new_sxml.encode("utf-8")
            elif item.filename == "xl/styles.xml":
                data = styles_xml.encode("utf-8")
            zi = zipfile.ZipInfo(item.filename, date_time=item.date_time)
            zi.compress_type = item.compress_type
            zi.external_attr = item.external_attr
            zout.writestr(zi, data)
    zin.close()
    return out.getvalue()


def validar_salida(original: bytes, nuevo: bytes, n_hojas: int):
    """Comprueba la integridad del .xlsm generado ANTES de devolverlo.

    Si algo no cuadra lanza 500: la ejecución de n8n se detiene y nunca llega
    al nodo que sobrescribe el archivo original en Drive. El peor caso pasa a
    ser 'no se procesó' en vez de 'se dañó el archivo del cliente'.
    """
    try:
        zo = zipfile.ZipFile(io.BytesIO(original))
    except Exception as e:
        raise HTTPException(500, f"El archivo original no es un ZIP válido: {e}")
    try:
        zn = zipfile.ZipFile(io.BytesIO(nuevo))
    except Exception as e:
        zo.close()
        raise HTTPException(500, f"El archivo generado no es un ZIP válido: {e}")

    try:
        faltantes = sorted(set(zo.namelist()) - set(zn.namelist()))
        if faltantes:
            raise HTTPException(
                500, f"El archivo generado perdió {len(faltantes)} parte(s) del original: "
                     f"{faltantes[:5]}")

        # Las macros tienen que salir idénticas byte a byte.
        if "xl/vbaProject.bin" in zo.namelist():
            if zo.read("xl/vbaProject.bin") != zn.read("xl/vbaProject.bin"):
                raise HTTPException(500, "Las macros (vbaProject.bin) cambiaron: se aborta")

        daniado = zn.testzip()
        if daniado:
            raise HTTPException(500, f"CRC inválido en el archivo generado: {daniado}")
    finally:
        zo.close()
        zn.close()

    # Última prueba: que abra de verdad y conserve todas sus hojas.
    try:
        wb2 = openpyxl.load_workbook(io.BytesIO(nuevo), keep_vba=True, data_only=False)
    except Exception as e:
        raise HTTPException(500, f"El archivo generado no se puede abrir: {e}")
    try:
        if len(wb2.worksheets) != n_hojas:
            raise HTTPException(
                500, f"El archivo generado tiene {len(wb2.worksheets)} hojas "
                     f"y el original {n_hojas}")
        find_sheet(wb2)      # lanza si la hoja objetivo desapareció
    finally:
        wb2.close()


@app.get("/health")
def health():
    return {"status": "ok", "service": "contaser-xlsx", "version": "1.9",
            "modulos": ["COMPRAS FC ELEC", "REPORTE DIAN"]}


@app.post("/process")
async def process(file: UploadFile = File(...), x_service_token: str = Header(default="")):
    if SERVICE_TOKEN and x_service_token != SERVICE_TOKEN:
        raise HTTPException(401, "Token de servicio inválido")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), keep_vba=True, data_only=False)
    except Exception as e:
        raise HTTPException(400, f"No se pudo abrir el archivo .xlsm: {e}")

    ws = find_sheet(wb)
    header_row = find_header_row(ws)
    colmap = map_columns(ws, header_row)
    anio, nit, nombre = parse_header(ws, header_row, file.filename)

    facturas = []
    last_row = header_row
    r = header_row + 1
    while r < header_row + 100000:
        nit_em = ws.cell(row=r, column=colmap["nit_emisor"]).value
        if END_MARKER in _norm(nit_em):
            break
        claves = [ws.cell(row=r, column=colmap[k]).value for k in ("nit_emisor", "valor_facturado", "cufe")]
        if all(v is None or str(v).strip() == "" for v in claves):
            break
        facturas.append({
            "nit_emisor": str(nit_em).strip() if nit_em is not None else "",
            "nombre_emisor": str(_celda(ws, r, colmap, "nombre_emisor") or "").strip(),
            "fecha_emision": to_iso(_celda(ws, r, colmap, "fecha_emision")),
            "valor_facturado": to_number(_celda(ws, r, colmap, "valor_facturado")),
            "notas_credito": to_number(_celda(ws, r, colmap, "notas_credito")),
            "notas_debito": to_number(_celda(ws, r, colmap, "notas_debito")),
            "valor_neto": to_number(_celda(ws, r, colmap, "valor_neto")),
            "valor_beneficio": to_number(_celda(ws, r, colmap, "valor_beneficio")),
            "medio_pago": str(_celda(ws, r, colmap, "medio_pago") or "").strip(),
            "num_factura": str(_celda(ws, r, colmap, "num_factura") or "").strip(),
            "cufe": str(_celda(ws, r, colmap, "cufe") or "").strip(),
        })
        last_row = r
        r += 1

    if not facturas:
        raise HTTPException(400, "No se encontraron facturas en la hoja COMPRAS FC ELEC")

    # ---- Totales ----
    t_fact = sum(f["valor_facturado"] for f in facturas)
    t_nc = sum(f["notas_credito"] for f in facturas)
    t_nd = sum(f["notas_debito"] for f in facturas)
    t_neto = sum(f["valor_neto"] for f in facturas)
    t_benef = sum(f["valor_beneficio"] for f in facturas)
    benef_1pct = round(t_benef * 0.01)

    # ---- Conteos por medio ----
    medios = {"electronico": {"count": 0, "valor": 0.0},
              "efectivo": {"count": 0, "valor": 0.0},
              "efectivo_otros": {"count": 0, "valor": 0.0},
              "otro": {"count": 0, "valor": 0.0}}
    for f in facturas:
        m = classify_medio(f["medio_pago"])
        medios[m]["count"] += 1
        medios[m]["valor"] += f["valor_facturado"]

    pct_elec = round((medios["electronico"]["valor"] / t_fact * 100), 1) if t_fact else 0
    pct_efec = round(100 - pct_elec, 1)

    # ---- Alertas ----
    vistos = {}
    duplicados = []
    for f in facturas:
        k = (f["nit_emisor"], f["num_factura"])
        if k in vistos:
            duplicados.append(f"{f['nit_emisor']} - {f['num_factura']}")
        vistos[k] = True
    alertas = {
        "facturas_en_cero": [f for f in facturas if f["valor_facturado"] == 0],
        "duplicadas": duplicados,
        "mayores_10m": [f for f in facturas if f["valor_facturado"] > 10_000_000],
        "efectivo_mayor_5m": [f for f in facturas
                              if classify_medio(f["medio_pago"]) in ("efectivo", "efectivo_otros")
                              and f["valor_facturado"] > 5_000_000],
        "desviacion_60_40": abs(pct_elec - 60) > 15,
    }

    # ---- Tabla final: base = Facturado - Notas Crédito ----
    benef_elec = sum(f["valor_beneficio"] for f in facturas
                     if classify_medio(f["medio_pago"]) == "electronico")
    base_fact_nc = t_fact - t_nc
    pct_benef = round(benef_elec / base_fact_nc * 100, 1) if base_fact_nc else 0.0

    def money(v):
        return int(round(v))

    tabla = [
        ("Facturado - Notas Crédito", money(base_fact_nc), "money"),
        ("Valor Susceptible Beneficio (Electrónicos)", money(benef_elec), "money"),
        ("60% (Facturado - NC)", money(0.6 * base_fact_nc), "money"),
        ("40% (Facturado - NC)", money(0.4 * base_fact_nc), "money"),
        # como fracción: Excel lo muestra 91,8% gracias al formato 0.0%
        ("% Beneficio / (Facturado - NC)",
         round(benef_elec / base_fact_nc, 4) if base_fact_nc else 0, "pct"),
    ]

    # ---- Módulo 2: comparación REPORTE DIAN (None si el libro no aplica) ----
    dian = procesar_modulo_dian(wb)

    # ---- Escritura quirúrgica sobre los bytes ORIGINALES ----
    # Los dos módulos escriben en la misma pasada de bytes: la salida del
    # Módulo 1 entra al Módulo 2, y Drive se actualiza UNA sola vez.
    hoja_titulo = ws.title
    n_hojas = len(wb.worksheets)
    nuevo = write_results_table(content, hoja_titulo, last_row, tabla)
    if dian:
        nuevo = write_dian_table(nuevo, dian["hoja_nueva"], dian.pop("_filas"))
    wb.close()                       # liberar antes de validar (abre otro libro)
    validar_salida(content, nuevo, n_hojas)
    archivo_b64 = base64.b64encode(nuevo).decode()

    return {
        "cliente": {"nit": str(nit).strip() if nit else "", "nombre": str(nombre).strip() if nombre else "",
                    "anio_gravable": int(to_number(anio)) if anio else None},
        "fecha_procesamiento": datetime.now().isoformat(timespec="seconds"),
        "total_facturas": len(facturas),
        "totales": {"valor_facturado": t_fact, "notas_credito": t_nc, "notas_debito": t_nd,
                    "valor_neto": t_neto, "valor_beneficio": t_benef},
        "beneficio_1pct": benef_1pct,
        "tabla_final": {
            "facturado_menos_nc": money(base_fact_nc),
            "beneficio_electronicos": money(benef_elec),
            "estimado_60": money(0.6 * base_fact_nc),
            "estimado_40": money(0.4 * base_fact_nc),
            "pct_beneficio_sobre_base": pct_benef,
        },
        "medios": medios,
        "pct_real": {"electronico": pct_elec, "efectivo": pct_efec},
        "alertas": alertas,
        "reporte_dian": dian,        # None si el libro no trae dos hojas REPORTE DIAN
        "facturas": facturas,
        "archivo_modificado_b64": archivo_b64,
    }


# ---------------------------------------------------------------------------
# Auditoría de riesgos (SOLO LECTURA: no modifica ni devuelve el archivo)
# ---------------------------------------------------------------------------

COLS_TABLA = {3, 4}          # C y D: donde write_results_table escribe
MAX_ITEMS = 200              # tope por lista para no devolver respuestas enormes

# Fórmulas: se buscan directamente y luego se ubica la celda contenedora
# mirando hacia atrás. Evita un lookahead por carácter sobre XML de megas.
_RE_F = re.compile(r"<f[^>]*>(.*?)</f>", re.DOTALL)
_RE_C_REF = re.compile(r'<c r="([A-Z]+\d+)"')

# Rango con fila explícita: D19, $D$19, D19:H400. El lookahead evita
# capturar nombres de función (SUM( ) y sufijos alfanuméricos.
_RE_RANGO = re.compile(
    r"(?:(?P<hoja>'[^']*'|[A-Za-z_][A-Za-z0-9_.]*)!)?"
    r"\$?(?P<c1>[A-Z]{1,3})\$?(?P<r1>\d+)"
    r"(?::\$?(?P<c2>[A-Z]{1,3})\$?(?P<r2>\d+))?"
    r"(?![A-Za-z0-9_(])")

# Columna completa: D:D, $C:$H  (el caso más peligroso para nuestra tabla)
_RE_COLFULL = re.compile(
    r"(?:(?P<hoja>'[^']*'|[A-Za-z_][A-Za-z0-9_.]*)!)?"
    r"\$?(?P<c1>[A-Z]{1,3})\$?:\$?(?P<c2>[A-Z]{1,3})\$?"
    r"(?![A-Za-z0-9_(])")


def _col_num(letras):
    n = 0
    for ch in str(letras).upper():
        if "A" <= ch <= "Z":
            n = n * 26 + (ord(ch) - 64)
    return n


def _rango_toca_cd(c1, c2):
    a = _col_num(c1)
    b = _col_num(c2) if c2 else a
    if a > b:
        a, b = b, a
    return any(a <= x <= b for x in COLS_TABLA)


def _mapa_hojas(z: zipfile.ZipFile):
    """[(nombre_hoja, ruta_xml), ...] para TODAS las hojas del libro."""
    wb_xml = z.read("xl/workbook.xml").decode("utf-8", errors="replace")
    rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8", errors="replace")
    rel_map = {}
    for rm in re.finditer(r"<Relationship\b[^>]*/?>", rels):
        tag = rm.group(0)
        rid = re.search(r'Id="([^"]*)"', tag)
        tgt = re.search(r'Target="([^"]*)"', tag)
        if rid and tgt:
            t = tgt.group(1)
            rel_map[rid.group(1)] = t[1:] if t.startswith("/") else "xl/" + t
    hojas = []
    for sm in re.finditer(r"<sheet\b[^>]*>", wb_xml):
        tag = sm.group(0)
        nm = re.search(r'name="([^"]*)"', tag)
        rid = re.search(r'r:id="([^"]*)"', tag)
        if nm and rid and rid.group(1) in rel_map:
            hojas.append((unescape(nm.group(1)), rel_map[rid.group(1)]))
    return hojas


def _rangos_de(formula, hoja_objetivo, es_hoja_objetivo):
    """Rangos de la fórmula que apuntan a la hoja objetivo.

    En la propia hoja objetivo, una referencia sin prefijo (D19:D400) apunta
    a esa hoja: es justo el caso más probable y más peligroso.
    """
    f = unescape(formula or "")
    obj = _norm_title(hoja_objetivo)
    encontrados = []
    for rx, tipo in ((_RE_COLFULL, "columna_completa"), (_RE_RANGO, "rango")):
        for m in rx.finditer(f):
            hoja = (m.group("hoja") or "").strip("'")
            propia = (not hoja and es_hoja_objetivo)
            if not propia and _norm_title(hoja) != obj:
                continue
            encontrados.append({
                "ref": m.group(0),
                "tipo": tipo,
                "toca_cd": _rango_toca_cd(m.group("c1"), m.group("c2")),
            })
    return encontrados


@app.post("/audit")
async def audit(file: UploadFile = File(...), x_service_token: str = Header(default="")):
    """Audita el .xlsm SIN modificarlo. Responde si la tabla de resultados
    puede contaminar fórmulas existentes y vuelca el header para diagnosticar
    el nombre de cliente vacío."""
    if SERVICE_TOKEN and x_service_token != SERVICE_TOKEN:
        raise HTTPException(401, "Token de servicio inválido")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), keep_vba=True, data_only=False)
    except Exception as e:
        raise HTTPException(400, f"No se pudo abrir el archivo .xlsm: {e}")

    ws = find_sheet(wb)
    header_row = find_header_row(ws)
    colmap = map_columns(ws, header_row)
    anio, nit, nombre = parse_header(ws, header_row, file.filename)

    # --- 1. Volcado del header: diagnostica el nombre de cliente vacío ---
    header_dump = []
    for r in range(1, header_row):
        celdas = {}
        for c in range(1, 9):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() != "":
                celdas[get_column_letter(c)] = str(v)[:120]
        if celdas:
            header_dump.append({"fila": r, "celdas": celdas})

    # --- 2. ¿Alguna columna de datos vive en C o D? ---
    columnas_datos = {campo: get_column_letter(c)
                      for campo, c in sorted(colmap.items(), key=lambda kv: kv[1])}
    conflicto_cd = sorted({v for v in columnas_datos.values()} & {"C", "D"})

    # --- 3. Extensión real de los datos (misma lógica que /process) ---
    last_row = header_row
    r = header_row + 1
    while r < header_row + 100000:
        nit_em = ws.cell(row=r, column=colmap["nit_emisor"]).value
        if END_MARKER in _norm(nit_em):
            break
        claves = [ws.cell(row=r, column=colmap[k]).value
                  for k in ("nit_emisor", "valor_facturado", "cufe")]
        if all(v is None or str(v).strip() == "" for v in claves):
            break
        last_row = r
        r += 1

    # --- 4. Qué hay hoy en las columnas C y D ---
    filas_muestra = sorted({x for x in (
        list(range(1, header_row + 1))
        + [header_row + 1, header_row + 2, last_row, last_row + 1, last_row + 2]
    ) if x >= 1})
    contenido_cd = []
    for r in filas_muestra:
        cv = ws.cell(row=r, column=3).value
        dv = ws.cell(row=r, column=4).value
        if cv is None and dv is None:
            continue
        contenido_cd.append({
            "fila": r,
            "C": str(cv)[:120] if cv is not None else None,
            "D": str(dv)[:120] if dv is not None else None,
        })

    # openpyxl ya no se necesita: liberar el libro antes del barrido del ZIP
    # (un .xlsm de 11 MB ocupa cientos de MB cargado en memoria).
    hoja_titulo = ws.title
    wb.close()

    zin = zipfile.ZipFile(io.BytesIO(content))
    hojas = _mapa_hojas(zin)
    obj_norm = _norm_title(hoja_titulo)

    # --- 5. Barrido de fórmulas en todas las hojas ---
    formulas_riesgo, formulas_referencian = [], []
    total_formulas = 0
    for nombre_hoja, ruta in hojas:
        try:
            sxml = zin.read(ruta).decode("utf-8", errors="replace")
        except KeyError:
            continue
        es_obj = _norm_title(nombre_hoja) == obj_norm
        for m in _RE_F.finditer(sxml):
            formula = m.group(1)
            if not formula.strip():
                continue
            total_formulas += 1
            # Celda contenedora = <c r="..."> más cercano hacia atrás
            celda = ""
            i = sxml.rfind('<c r="', 0, m.start())
            if i != -1:
                cm = _RE_C_REF.match(sxml, i)
                if cm:
                    celda = cm.group(1)
            rel = _rangos_de(formula, hoja_titulo, es_obj)
            if not rel:
                continue
            peligrosos = [rg["ref"] for rg in rel if rg["toca_cd"]]
            item = {
                "hoja": nombre_hoja,
                "celda": celda,
                "formula": unescape(formula)[:300],
                "rangos_sobre_hoja_objetivo": [rg["ref"] for rg in rel][:20],
                "rangos_que_tocan_C_o_D": peligrosos[:20],
            }
            if len(formulas_referencian) < MAX_ITEMS:
                formulas_referencian.append(item)
            if peligrosos and len(formulas_riesgo) < MAX_ITEMS:
                formulas_riesgo.append(item)

    # --- 6. Nombres definidos que apunten a la hoja ---
    wb_xml = zin.read("xl/workbook.xml").decode("utf-8", errors="replace")
    nombres_definidos = []
    for m in re.finditer(r"<definedName\b([^>]*)>(.*?)</definedName>", wb_xml, re.DOTALL):
        nm = re.search(r'name="([^"]*)"', m.group(1))
        crudo = m.group(2)          # _rangos_de hace su propio unescape
        rel = _rangos_de(crudo, hoja_titulo, False)
        if rel:
            nombres_definidos.append({
                "nombre": nm.group(1) if nm else "",
                "refiere_a": unescape(crudo)[:300],
                "toca_C_o_D": any(rg["toca_cd"] for rg in rel),
            })

    # --- 7. Rastro de tablas anteriores (idempotencia) ---
    marcador = []
    for nombre_hoja, ruta in hojas:
        try:
            sxml = zin.read(ruta).decode("utf-8", errors="replace")
        except KeyError:
            continue
        if MARCADOR_TABLA not in sxml:
            continue
        for mm in re.finditer(r'<row r="(\d+)"[^>]*?(?:/>|>.*?</row>)', sxml, re.DOTALL):
            if MARCADOR_TABLA in mm.group(0):
                marcador.append({"hoja": nombre_hoja, "fila": int(mm.group(1))})

    # Si Excel reguardó el archivo, el texto pudo migrar a sharedStrings y la
    # idempotencia de write_results_table dejaría de encontrarlo -> tabla doble.
    marcador_shared = False
    try:
        ss = zin.read("xl/sharedStrings.xml").decode("utf-8", errors="replace")
        marcador_shared = MARCADOR_TABLA in ss
    except KeyError:
        pass

    # --- 8. Dónde caería la tabla ---
    sheet_path = _sheet_xml_path(zin, hoja_titulo)
    sxml_obj = zin.read(sheet_path).decode("utf-8", errors="replace")
    existentes = [int(x) for x in re.findall(r'<row r="(\d+)"', sxml_obj)]
    max_existente = max(existentes) if existentes else last_row
    start = max(last_row + 3, max_existente + 2)

    zin.close()

    n_riesgo = len(formulas_riesgo)
    hay_riesgo = bool(n_riesgo or conflicto_cd
                      or any(n["toca_C_o_D"] for n in nombres_definidos))

    return {
        "archivo": file.filename,
        "version_auditoria": "1.8",
        "hoja_objetivo": hoja_titulo,
        "total_hojas": len(hojas),
        "fila_encabezados": header_row,
        "ultima_fila_datos": last_row,
        "total_facturas": max(0, last_row - header_row),
        "header_parseado": {
            "anio_gravable": str(anio) if anio is not None else None,
            "nit": str(nit).strip() if nit else "",
            "nombre": str(nombre).strip() if nombre else "",
            "nombre_vacio": not (nombre and str(nombre).strip()),
        },
        "header_dump": header_dump,
        "columnas_datos": columnas_datos,
        "columnas_datos_en_C_o_D": conflicto_cd,
        "contenido_actual_C_D": contenido_cd,
        "zona_escritura": {
            "fila_inicio_tabla": start,
            "filas_que_ocupa": 7,
            "columnas": ["C", "D"],
            "max_fila_existente": max_existente,
        },
        "formulas_que_tocan_C_o_D": formulas_riesgo,
        "formulas_que_referencian_la_hoja": formulas_referencian,
        "nombres_definidos": nombres_definidos,
        "tabla_previa_detectada": marcador,
        "marcador_en_sharedStrings": marcador_shared,
        "resumen": {
            "total_formulas_del_libro": total_formulas,
            "formulas_que_referencian_la_hoja": len(formulas_referencian),
            "formulas_en_riesgo": n_riesgo,
            "columnas_de_datos_invadidas": conflicto_cd,
            "listas_truncadas": (len(formulas_referencian) >= MAX_ITEMS
                                 or n_riesgo >= MAX_ITEMS),
            "veredicto": "RIESGO DETECTADO" if hay_riesgo else "SIN RIESGO DETECTADO",
        },
    }


# ---------------------------------------------------------------------------
# MÓDULO 2 — Comparación REPORTE DIAN
# ---------------------------------------------------------------------------
# La hoja base es la que Contaser YA trabajó: trae anotaciones propias (los
# cinco 'Tope N', textos de 'Uso declaración' más largos y dos columnas extra
# del consultante). La hoja nueva es un export crudo. Por eso las diferencias
# de TEXTO son ruido y el emparejamiento se hace por NIT + Valor.

DIAN_PREFIJO = "reporte dian"
# Sin tildes a propósito: este texto es además el marcador de idempotencia y
# se busca dentro del XML, así que conviene mantenerlo lo más simple posible.
MARCADOR_DIAN = "COMPARACION REPORTE DIAN"

_RE_RENGLON = re.compile(r"\bR\s?(\d{1,3})\b")
_RE_TOPE = re.compile(r"\btope\s*(\d)\b", re.I)


def _clave_fecha(s):
    """Clave ordenable a partir de una fecha ('2025-08-19 09:12:51' -> '20250819091251')."""
    return re.sub(r"[^0-9]", "", str(s or ""))[:14].ljust(14, "0")


def _header_row_dian(ws):
    """Fila de encabezados: la única que trae NIT, Detalle y Valor a la vez."""
    for r in range(1, MAX_FILA_HEADER):
        etiquetas = {_norm_label(ws.cell(row=r, column=c).value)
                     for c in range(1, MAX_COL_HEADER + 1)}
        if {"nit", "detalle", "valor"} <= etiquetas:
            return r
    raise HTTPException(400, "REPORTE DIAN: no se encontró la fila de encabezados "
                             "(se esperaban NIT, Detalle y Valor en la misma fila)")


def _fecha_reporte_dian(ws, header_row):
    """'Fecha Reporte' del bloque superior de la hoja."""
    for r in range(1, header_row):
        for c in range(1, MAX_COL_HEADER + 1):
            if _norm_label(ws.cell(row=r, column=c).value).startswith("fecha reporte"):
                for cc in range(c + 1, MAX_COL_HEADER + 1):
                    v = ws.cell(row=r, column=cc).value
                    if v is not None and str(v).strip():
                        return v.isoformat() if isinstance(v, (datetime, date)) else str(v).strip()
    return ""


def _map_columns_dian(ws, header_row):
    """Mapea las columnas por etiqueta. GANA LA PRIMERA APARICIÓN: la hoja
    base repite 'NIT' y 'Nombre' para el consultante y no queremos esas."""
    colmap = {}
    encontradas = []
    for c in range(1, MAX_COL_HEADER + 1):
        crudo = ws.cell(row=header_row, column=c).value
        lab = _norm_label(crudo)
        if not lab:
            continue
        encontradas.append(f"{get_column_letter(c)}={crudo}")
        if lab == "nit":
            colmap.setdefault("nit", c)
        elif lab.startswith("nombre"):
            colmap.setdefault("tercero", c)
        elif lab == "detalle":
            colmap.setdefault("detalle", c)
        elif lab == "valor":
            colmap.setdefault("valor", c)
        elif lab.startswith("uso declaracion"):
            colmap.setdefault("uso", c)
        elif lab.startswith("informacion adicional"):
            colmap.setdefault("info", c)
    faltan = [k for k in ("nit", "detalle", "valor") if k not in colmap]
    if faltan:
        raise HTTPException(400, f"REPORTE DIAN ({ws.title}): faltan columnas {faltan}. "
                                 f"Encabezados en la fila {header_row}: {encontradas[:20]}")
    return colmap


def _fila_marcador(ws, marcador):
    """Fila donde empieza un cuadro escrito por nosotros, o None."""
    for r in range(1, (ws.max_row or 1) + 1):
        for c in range(1, 10):
            v = ws.cell(row=r, column=c).value
            if v is not None and marcador in str(v):
                return r
    return None


def _leer_registros_dian(ws, header_row, colmap):
    """Registros de la hoja. Descarta filas sin NIT (ahí caen los 'Tope N')
    y se detiene antes de un cuadro de comparación previo, para no leer
    nuestra propia salida como si fueran registros de la DIAN."""
    tope = _fila_marcador(ws, MARCADOR_DIAN) or ((ws.max_row or header_row) + 1)
    registros = []
    for r in range(header_row + 1, tope):
        nit = _celda(ws, r, colmap, "nit")
        nit_s = str(nit).strip() if nit is not None else ""
        if not nit_s:
            continue
        registros.append({
            "nit": nit_s,
            "tercero": str(_celda(ws, r, colmap, "tercero") or "").strip(),
            "detalle": str(_celda(ws, r, colmap, "detalle") or "").strip(),
            "valor": to_number(_celda(ws, r, colmap, "valor")),
            "uso": str(_celda(ws, r, colmap, "uso") or "").strip(),
            "fila": r,
        })
    return registros


def _pref(s):
    return _norm_label(s)[:40]


def comparar_registros_dian(base, nueva):
    """Reconciliación en tres pasadas. Devuelve (sin_cambio, cambiados,
    nuevos, desaparecidos, ambiguos)."""
    usados = set()
    sin_cambio, cambiados, nuevos, ambiguos = [], [], [], []

    # Pasada 1 — NIT + Valor exacto: el registro no cambió.
    idx1 = {}
    for i, reg in enumerate(base):
        idx1.setdefault((reg["nit"], round(reg["valor"], 2)), []).append(i)
    resto_nueva = []
    for reg in nueva:
        libres = [i for i in idx1.get((reg["nit"], round(reg["valor"], 2)), [])
                  if i not in usados]
        if libres:
            usados.add(libres[0])
            sin_cambio.append(reg)
            if len(libres) > 1:
                ambiguos.append({**reg, "motivo": "varios registros del mismo NIT con igual valor"})
        else:
            resto_nueva.append(reg)

    # Pasada 2 — NIT + prefijo del Detalle: cambió el valor.
    for reg in resto_nueva:
        p = _pref(reg["detalle"])
        elegido = None
        for i, b in enumerate(base):
            if i in usados or b["nit"] != reg["nit"]:
                continue
            pb = _pref(b["detalle"])
            if pb == p or pb.startswith(p) or p.startswith(pb):
                elegido = i
                break
        if elegido is None:
            nuevos.append(reg)
        else:
            usados.add(elegido)
            ant = base[elegido]["valor"]
            cambiados.append({**reg, "valor_anterior": ant,
                              "diferencia": round(reg["valor"] - ant, 2)})

    # Pasada 3 — lo que quedó sin emparejar en la base desapareció.
    desaparecidos = [b for i, b in enumerate(base) if i not in usados]
    return sin_cambio, cambiados, nuevos, desaparecidos, ambiguos


def impacto_dian(nuevos, cambiados):
    """Agrupa el impacto por renglón de la declaración (R29, R74...) y por
    Tope. Se mantienen separados a propósito: un mismo registro puede afectar
    un renglón y un tope, y sumarlos juntos daría un total engañoso."""
    renglones, topes, sin_clasificar = {}, {}, {"valor": 0.0, "registros": 0}

    def sumar(acum, clave, descripcion, monto):
        e = acum.setdefault(clave, {"codigo": clave, "descripcion": descripcion,
                                    "valor": 0.0, "registros": 0})
        e["valor"] += monto
        e["registros"] += 1

    for reg, monto in ([(r, r["valor"]) for r in nuevos]
                       + [(r, r["diferencia"]) for r in cambiados]):
        uso = reg.get("uso") or ""
        clasificado = False
        for m in _RE_RENGLON.finditer(uso):
            desc = uso[m.end():].split("|")[0].strip(" -–|") or uso[:60]
            sumar(renglones, "R" + m.group(1), desc[:60], monto)
            clasificado = True
        for m in _RE_TOPE.finditer(uso):
            sumar(topes, "Tope " + m.group(1), uso[m.end():].split("|")[0].strip(" .-–|")[:60], monto)
            clasificado = True
        if not clasificado:
            sin_clasificar["valor"] += monto
            sin_clasificar["registros"] += 1

    orden = lambda d: sorted(d.values(), key=lambda e: -abs(e["valor"]))
    return {"renglones": orden(renglones), "topes": orden(topes),
            "sin_clasificar": sin_clasificar}


def construir_cuadro_dian(base_m, nueva_m, sin_cambio, cambiados, nuevos,
                          desaparecidos, impacto):
    """Arma las filas del cuadro. Cada celda es (columna, valor, estilo)."""
    B, C, D, E, F = "B", "C", "D", "E", "F"
    filas = []
    suma = lambda regs: int(round(sum(r["valor"] for r in regs)))

    filas.append([(B, MARCADOR_DIAN, "titulo"), (C, "", "titulo"),
                  (D, "", "titulo"), (E, "", "titulo"), (F, "", "titulo")])
    filas.append([(B, "Hoja base (ya trabajada)", "label"),
                  (C, f"{base_m['hoja']}  |  {base_m['fecha']}", None)])
    filas.append([(B, "Hoja nueva", "label"),
                  (C, f"{nueva_m['hoja']}  |  {nueva_m['fecha']}", None)])
    filas.append([])

    filas.append([(B, "Concepto", "header"), (C, "Cantidad", "header"),
                  (D, "Valor", "header")])
    filas.append([(B, "Registros sin cambio", "label"),
                  (C, len(sin_cambio), "money"), (D, suma(sin_cambio), "money")])
    filas.append([(B, "Registros NUEVOS", "label"),
                  (C, len(nuevos), "money"), (D, suma(nuevos), "money")])
    filas.append([(B, "Registros con VALOR CAMBIADO", "label"),
                  (C, len(cambiados), "money"),
                  (D, int(round(sum(r["diferencia"] for r in cambiados))), "money")])
    filas.append([(B, "Registros DESAPARECIDOS", "label"),
                  (C, len(desaparecidos), "money"), (D, suma(desaparecidos), "money")])
    filas.append([(B, "Total hoja base", "label"),
                  (C, len(sin_cambio) + len(cambiados) + len(desaparecidos), "money"),
                  (D, suma(sin_cambio) + suma(desaparecidos)
                      + int(round(sum(r["valor_anterior"] for r in cambiados))), "money")])
    filas.append([(B, "Total hoja nueva", "label"),
                  (C, len(sin_cambio) + len(cambiados) + len(nuevos), "money"),
                  (D, suma(sin_cambio) + suma(cambiados) + suma(nuevos), "money")])
    filas.append([])

    if impacto["renglones"] or impacto["topes"]:
        filas.append([(B, "IMPACTO EN LA DECLARACION (solo nuevos y cambiados)", "titulo"),
                      (C, "", "titulo"), (D, "", "titulo"), (E, "", "titulo"), (F, "", "titulo")])
        filas.append([(B, "Renglon / Tope", "header"), (C, "Descripcion", "header"),
                      (D, "Registros", "header"), (E, "Valor", "header")])
        for e in impacto["renglones"] + impacto["topes"]:
            filas.append([(B, e["codigo"], "label"), (C, e["descripcion"], None),
                          (D, e["registros"], "money"), (E, int(round(e["valor"])), "money")])
        if impacto["sin_clasificar"]["registros"]:
            filas.append([(B, "SIN CLASIFICAR", "label"),
                          (C, "sin codigo de renglon en 'Uso declaracion Sugerida'", None),
                          (D, impacto["sin_clasificar"]["registros"], "money"),
                          (E, int(round(impacto["sin_clasificar"]["valor"])), "money")])
        filas.append([])

    def bloque_detalle(titulo, regs, con_anterior=False):
        if not regs:
            return
        filas.append([(B, titulo, "titulo"), (C, "", "titulo"), (D, "", "titulo"),
                      (E, "", "titulo"), (F, "", "titulo")])
        cab = [(B, "NIT", "header"), (C, "Tercero", "header"),
               (D, "Detalle", "header"), (E, "Valor", "header")]
        if con_anterior:
            cab.append((F, "Valor anterior", "header"))
        filas.append(cab)
        for r in regs:
            fila = [(B, r["nit"], None), (C, r["tercero"], None),
                    (D, r["detalle"], None), (E, int(round(r["valor"])), "money")]
            if con_anterior:
                fila.append((F, int(round(r["valor_anterior"])), "money"))
            filas.append(fila)
        filas.append([])

    bloque_detalle("REGISTROS NUEVOS", nuevos)
    bloque_detalle("REGISTROS CON VALOR CAMBIADO", cambiados, con_anterior=True)
    bloque_detalle("REGISTROS DESAPARECIDOS", desaparecidos)
    return filas


def write_dian_table(content: bytes, sheet_title: str, filas: list) -> bytes:
    """Escribe el cuadro al final de la hoja nueva, con la misma cirugía de
    ZIP/XML del Módulo 1: solo se tocan el XML de esa hoja y styles.xml.

    OJO: por idempotencia se borra todo lo que haya desde la fila del
    marcador hacia abajo. El cuadro siempre va al final, así que no debería
    haber nada más ahí; no escribas notas manuales debajo del cuadro.
    """
    zin = zipfile.ZipFile(io.BytesIO(content))
    sheet_path = _sheet_xml_path(zin, sheet_title)
    sxml = zin.read(sheet_path).decode("utf-8")
    styles_xml, st = _augment_styles(zin.read("xl/styles.xml").decode("utf-8"))

    row_re = re.compile(r'<row r="(\d+)"[^>]*?(?:/>|>.*?</row>)', re.DOTALL)
    sst_idx = _indices_sharedstrings(zin, MARCADOR_DIAN)
    cell_s_re = re.compile(r'<c\b[^>]*\bt="s"[^>]*>\s*<v>(\d+)</v>')

    title_row = None
    for m in row_re.finditer(sxml):
        bloque = m.group(0)
        if MARCADOR_DIAN in bloque or (
                sst_idx and any(int(cm.group(1)) in sst_idx
                                for cm in cell_s_re.finditer(bloque))):
            title_row = int(m.group(1))
            break
    if title_row is not None:
        sxml = row_re.sub(lambda m: "" if int(m.group(1)) >= title_row else m.group(0), sxml)

    existentes = [int(x) for x in re.findall(r'<row r="(\d+)"', sxml)]
    start = (max(existentes) if existentes else 1) + 2

    def celda(col, r, valor, kind):
        if kind in ("money", "pct"):
            return '<c r="%s%d" s="%d"><v>%s</v></c>' % (col, r, st[kind], valor)
        s_attr = ' s="%d"' % st[kind] if kind in st else ""
        return ('<c r="%s%d"%s t="inlineStr"><is><t xml:space="preserve">%s</t></is></c>'
                % (col, r, s_attr, escape(str(valor))))

    rows_xml = []
    for i, fila in enumerate(filas):
        r = start + i
        cuerpo = "".join(celda(col, r, val, kind) for col, val, kind in fila)
        rows_xml.append('<row r="%d">%s</row>' % (r, cuerpo))

    idx = sxml.rfind("</sheetData>")
    if idx == -1:
        raise HTTPException(500, "XML de hoja sin </sheetData>; estructura inesperada")
    new_sxml = sxml[:idx] + "".join(rows_xml) + sxml[idx:]

    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == sheet_path:
                data = new_sxml.encode("utf-8")
            elif item.filename == "xl/styles.xml":
                data = styles_xml.encode("utf-8")
            zi = zipfile.ZipInfo(item.filename, date_time=item.date_time)
            zi.compress_type = item.compress_type
            zi.external_attr = item.external_attr
            zout.writestr(zi, data)
    zin.close()
    return out.getvalue()


def procesar_modulo_dian(wb):
    """Devuelve el resultado del Módulo 2, o None si no aplica (menos de dos
    hojas REPORTE DIAN con estructura de reporte)."""
    hojas = []
    for ws in wb.worksheets:
        if not _norm_title(ws.title).startswith(DIAN_PREFIJO):
            continue
        try:
            hr = _header_row_dian(ws)
        except HTTPException:
            continue          # se llama REPORTE DIAN pero no tiene la estructura
        fecha = _fecha_reporte_dian(ws, hr)
        hojas.append({"ws": ws, "hoja": ws.title, "header_row": hr,
                      "fecha": fecha, "clave": _clave_fecha(fecha)})
    if len(hojas) < 2:
        return None

    hojas.sort(key=lambda h: h["clave"])
    base_m, nueva_m = hojas[0], hojas[-1]

    base = _leer_registros_dian(base_m["ws"], base_m["header_row"],
                                _map_columns_dian(base_m["ws"], base_m["header_row"]))
    nueva = _leer_registros_dian(nueva_m["ws"], nueva_m["header_row"],
                                 _map_columns_dian(nueva_m["ws"], nueva_m["header_row"]))

    sin_cambio, cambiados, nuevos, desaparecidos, ambiguos = comparar_registros_dian(base, nueva)
    impacto = impacto_dian(nuevos, cambiados)
    filas = construir_cuadro_dian(base_m, nueva_m, sin_cambio, cambiados,
                                  nuevos, desaparecidos, impacto)
    limpiar = lambda regs: [{k: v for k, v in r.items() if k != "fila"} for r in regs]
    return {
        "hoja_base": base_m["hoja"],
        "fecha_base": base_m["fecha"],
        "hoja_nueva": nueva_m["hoja"],
        "fecha_nueva": nueva_m["fecha"],
        "hojas_detectadas": [{"hoja": h["hoja"], "fecha": h["fecha"]} for h in hojas],
        "conteos": {"base": len(base), "nueva": len(nueva),
                    "sin_cambio": len(sin_cambio), "cambiados": len(cambiados),
                    "nuevos": len(nuevos), "desaparecidos": len(desaparecidos)},
        "impacto": impacto,
        "nuevos": limpiar(nuevos),
        "cambiados": limpiar(cambiados),
        "desaparecidos": limpiar(desaparecidos),
        "ambiguos": limpiar(ambiguos),
        "filas_cuadro": len(filas),
        "_filas": filas,
    }


@app.post("/dian-preview")
async def dian_preview(file: UploadFile = File(...), x_service_token: str = Header(default="")):
    """Corre el Módulo 2 y devuelve el resultado SIN escribir nada.

    Permite validar la comparación (cuántos nuevos, cuánto impacto, si el
    emparejamiento tiene sentido) antes de dejar que toque un archivo real.
    """
    if SERVICE_TOKEN and x_service_token != SERVICE_TOKEN:
        raise HTTPException(401, "Token de servicio inválido")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), keep_vba=True, data_only=False)
    except Exception as e:
        raise HTTPException(400, f"No se pudo abrir el archivo .xlsm: {e}")

    candidatas = [w.title for w in wb.worksheets
                  if _norm_title(w.title).startswith(DIAN_PREFIJO)]
    try:
        dian = procesar_modulo_dian(wb)
    finally:
        wb.close()

    if dian is None:
        return {"aplica": False,
                "hojas_con_nombre_reporte_dian": candidatas,
                "motivo": "Se necesitan dos hojas REPORTE DIAN con estructura de reporte "
                          "(fila de encabezados con NIT, Detalle y Valor)"}
    dian.pop("_filas", None)
    return {"aplica": True, "archivo": file.filename, **dian}


# ---------------------------------------------------------------------------
# Inspección genérica de hojas (SOLO LECTURA)
# ---------------------------------------------------------------------------

def _buscar_hoja(wb, nombre):
    """Localiza una hoja por nombre con tolerancia (igual que find_sheet,
    pero para cualquier hoja, no solo COMPRAS FC ELEC)."""
    objetivo = _norm_title(nombre)
    for w in wb.worksheets:
        if _norm_title(w.title) == objetivo:
            return w
    for w in wb.worksheets:
        if objetivo and objetivo in _norm_title(w.title):
            return w
    compacto = objetivo.replace(" ", "")
    for w in wb.worksheets:
        if compacto and compacto in _norm_title(w.title).replace(" ", ""):
            return w
    return None


@app.post("/inspect")
async def inspect(file: UploadFile = File(...),
                  hoja: str = Form(default=""),
                  filas: int = Form(default=25),
                  x_service_token: str = Header(default="")):
    """Devuelve la estructura de una hoja SIN modificar el archivo.

    Sirve para diseñar los módulos siguientes (REPORTE DIAN, PATRIMONIO,
    CED.1 GENERAL...) sobre datos reales en vez de suposiciones. Si no se
    indica `hoja`, lista todas las hojas del libro con sus dimensiones.
    """
    if SERVICE_TOKEN and x_service_token != SERVICE_TOKEN:
        raise HTTPException(401, "Token de servicio inválido")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), keep_vba=True, data_only=False)
    except Exception as e:
        raise HTTPException(400, f"No se pudo abrir el archivo .xlsm: {e}")

    hojas = [{"nombre": w.title, "max_fila": w.max_row, "max_columna": w.max_column}
             for w in wb.worksheets]

    if not (hoja or "").strip():
        wb.close()
        return {"archivo": file.filename, "total_hojas": len(hojas), "hojas": hojas}

    ws = _buscar_hoja(wb, hoja)
    if ws is None:
        nombres = [h["nombre"] for h in hojas]
        wb.close()
        raise HTTPException(400, f"No se encontró la hoja {hoja!r}. Hojas del libro: {nombres}")

    tope_filas = max(1, min(int(filas or 25), 200))
    max_col = min(ws.max_column or 1, MAX_COL_HEADER)
    muestra = []
    for r in range(1, min(ws.max_row or 1, 2000) + 1):
        celdas = {}
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() != "":
                celdas[get_column_letter(c)] = str(v)[:120]
        if celdas:
            muestra.append({"fila": r, "celdas": celdas})
            if len(muestra) >= tope_filas:
                break

    resultado = {
        "archivo": file.filename,
        "hoja": ws.title,
        "dimensiones": {"max_fila": ws.max_row, "max_columna": ws.max_column,
                        "ultima_columna_letra": get_column_letter(ws.max_column or 1)},
        "filas_mostradas": len(muestra),
        "muestra": muestra,
        "hojas_del_libro": [h["nombre"] for h in hojas],
    }
    wb.close()
    return resultado
